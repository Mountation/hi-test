from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from ..models import EvaluationSet, Corpus
import json
import openpyxl
from django.db import transaction


def api_list_datasets(request):
    sets = EvaluationSet.objects.filter(status='active').order_by('-created_at')
    data = [{'id': s.id, 'name': s.name, 'description': s.description, 'created_at': s.created_at.isoformat(), 'corpus_count': s.get_corpora_count()} for s in sets]
    return JsonResponse({'status': 'success', 'data': data})


@csrf_exempt
def api_create_dataset(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    if not request.FILES.get('excel_file'):
        return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

    excel_file = request.FILES['excel_file']
    name = request.POST.get('evaluation_set_name', '').strip()
    desc = request.POST.get('evaluation_set_desc', '')

    if not name:
        return JsonResponse({'status': 'error', 'message': 'Name required'}, status=400)

    if EvaluationSet.objects.filter(name=name).exists():
        return JsonResponse({'status': 'error', 'message': 'Name exists'}, status=400)

    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row or len(first_row) < 1:
            workbook.close()
            return JsonResponse({'status': 'error', 'message': 'Invalid Excel format'}, status=400)

        evaluation_set = EvaluationSet.objects.create(name=name, description=desc, status='active')

        batch_size = 1000
        objs = []
        success_count = 0
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    content = str(row[0])
                    expected_response = str(row[1]) if len(row) > 1 and row[1] is not None else None
                    intent = str(row[2]) if len(row) > 2 and row[2] is not None else None
                    objs.append(Corpus(evaluation_set=evaluation_set, content=content, expected_response=expected_response, intent=intent))
                    if len(objs) >= batch_size:
                        Corpus.objects.bulk_create(objs, batch_size=batch_size)
                        success_count += len(objs)
                        objs = []
            if objs:
                Corpus.objects.bulk_create(objs, batch_size=batch_size)
                success_count += len(objs)
        workbook.close()

        return JsonResponse({'status': 'success', 'message': 'Created', 'created': success_count, 'id': evaluation_set.id})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def api_dataset_detail(request, dataset_id):
    evaluation_set = get_object_or_404(EvaluationSet, id=dataset_id)
    corpora_qs = Corpus.objects.filter(evaluation_set=evaluation_set).order_by('created_at')
    page = request.GET.get('page', 1)
    paginator = Paginator(corpora_qs, 10)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    items = []
    for c in page_obj.object_list:
        items.append({'id': c.id, 'content': c.content, 'expected_response': c.expected_response, 'intent': c.intent, 'created_at': c.created_at.isoformat()})

    return JsonResponse({'status': 'success', 'data': items, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})


@csrf_exempt
def api_delete_dataset(request, dataset_id):
    if request.method != 'POST' and request.method != 'DELETE':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    evaluation_set = get_object_or_404(EvaluationSet, id=dataset_id)
    count = evaluation_set.get_corpora_count()
    evaluation_set.delete()
    return JsonResponse({'status': 'success', 'deleted': count})
