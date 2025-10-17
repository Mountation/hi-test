# myapp/views/dataset.py
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.db import transaction
from ..models import EvaluationSet, Corpus
import openpyxl
from collections import Counter
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def list_datasets(request):
    """
    列出所有评测集
    """
    evaluation_sets = EvaluationSet.objects.filter(status='active')
    return render(request, 'dataset/list.html', {'evaluation_sets': evaluation_sets})

def create_dataset(request):
    """
    创建评测集：上传Excel文件并将其内容存储到数据库
    """
    # 处理GET请求，显示创建页面
    if request.method != 'POST':
        return render(request, 'dataset/create.html')
    
    # 处理POST请求
    if not request.FILES.get('excel_file'):
        return JsonResponse({
            'status': 'error',
            'message': '请选择要上传的Excel文件'
        })
    
    excel_file = request.FILES['excel_file']
    evaluation_set_name = request.POST.get('evaluation_set_name', '').strip()
    evaluation_set_desc = request.POST.get('evaluation_set_desc', 'Uploaded from Excel file')
    
    # 检查评测集名称
    if not evaluation_set_name:
        return JsonResponse({
            'status': 'error',
            'message': '请输入评测集名称'
        })
    
    # 检查评测集名称是否已存在
    if EvaluationSet.objects.filter(name=evaluation_set_name).exists():
        return JsonResponse({
            'status': 'error',
            'message': f'评测集名称 "{evaluation_set_name}" 已存在，请使用其他名称'
        })
    
    try:
        # 使用 openpyxl 的只读模式按行流式读取，避免一次性把整个文件载入内存
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # 检查Excel格式（只检查第一行是否至少有一列）
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row or len(first_row) < 1:
            workbook.close()
            return JsonResponse({
                'status': 'error',
                'message': 'Excel文件格式不正确，至少需要一列数据'
            })

        # 创建评测集
        evaluation_set = EvaluationSet.objects.create(
            name=evaluation_set_name,
            description=evaluation_set_desc,
            status='active'
        )

        # 按批次构建 Corpus 实例并使用 bulk_create 提交，显著减少数据库交互次数
        batch_size = 1000
        objs = []
        success_count = 0

        # 使用事务包裹批量插入，保证一致性并提升性能
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    content = str(row[0])
                    expected_response = str(row[1]) if len(row) > 1 and row[1] is not None else None
                    intent = str(row[2]) if len(row) > 2 and row[2] is not None else None

                    objs.append(Corpus(
                        evaluation_set=evaluation_set,
                        content=content,
                        expected_response=expected_response,
                        intent=intent
                    ))

                    # 达到批次大小时写入数据库
                    if len(objs) >= batch_size:
                        Corpus.objects.bulk_create(objs, batch_size=batch_size)
                        success_count += len(objs)
                        objs = []

            # 插入剩余未满一批的数据
            if objs:
                Corpus.objects.bulk_create(objs, batch_size=batch_size)
                success_count += len(objs)

        # 关闭只读工作簿释放文件句柄
        workbook.close()
        
        return JsonResponse({
            'status': 'success',
            'message': f'成功创建评测集 "{evaluation_set_name}"，导入了 {success_count} 条语料记录'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'上传文件失败: {str(e)}'
        })

def delete_dataset(request, dataset_id):
    """
    删除评测集及其关联的所有语料
    """
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': '无效的请求方法'
        })
    
    # 获取评测集对象
    evaluation_set = get_object_or_404(EvaluationSet, id=dataset_id)

    try:
        # 获取关联的语料数量，用户返回信息（在删除前统计）
        corpus_qs = Corpus.objects.filter(evaluation_set=evaluation_set)
        corpus_count = corpus_qs.count()

        evaluation_set_name = evaluation_set.name

        # 使用事务执行批量删除：先以 QuerySet 方式删除所有语料（在 DB 层一次性删除，避免逐条触发 ORM delete）
        # 然后删除评测集本身
        with transaction.atomic():
            if corpus_count > 0:
                corpus_qs.delete()

            # 删除评测集（单条删除，开销很小）
            EvaluationSet.objects.filter(id=evaluation_set.id).delete()

        return JsonResponse({
            'status': 'success',
            'message': f'成功删除评测集 "{evaluation_set_name}"，删除了 {corpus_count} 条语料记录'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'删除评测集失败: {str(e)}'
        })
    
def view_dataset(request, dataset_id):
    """
    查看评测集详情，包括其中的所有语料
    """
    # 获取评测集对象
    evaluation_set = get_object_or_404(EvaluationSet, id=dataset_id)
    
    # 获取该评测集下的所有语料（作为 QuerySet）
    corpora_qs = Corpus.objects.filter(evaluation_set=evaluation_set).order_by('created_at')

    # 分页：每页显示 10 条
    page = request.GET.get('page', 1)
    paginator = Paginator(corpora_qs, 10)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    corpora_count = corpora_qs.count()

    context = {
        'evaluation_set': evaluation_set,
        # 将 page_obj 传入模板以便渲染分页控件，corpora 在模板中仍可用于迭代
        'corpora': page_obj,
        'page_obj': page_obj,
        'corpora_count': corpora_count,
        'start_index': page_obj.start_index(),
        'end_index': page_obj.end_index(),
    }
    
    return render(request, 'dataset/view.html', context)

def run_dataset_page(request):
    """
    运行评测集页面
    """
    # 获取所有活跃的评测集
    evaluation_sets = EvaluationSet.objects.filter(status='active')
    
    # 为每个评测集预处理意图数据
    for eval_set in evaluation_sets:
        # 获取该评测集下的所有非空意图
        intents = Corpus.objects.filter(
            evaluation_set=eval_set
        ).exclude(
            intent__isnull=True
        ).exclude(
            intent__exact=''
        ).values_list('intent', flat=True)
        
        # 统计意图出现次数并获取前5个最常见的意图
        intent_counter = Counter(intents)
        eval_set.top_intents = [intent for intent, count in intent_counter.most_common(5)]
        eval_set.intent_count = len(intent_counter)
    
    return render(request, 'dataset/run.html', {'evaluation_sets': evaluation_sets})