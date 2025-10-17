"""Utility helpers for CaseAI app.

Contains helpers for parsing Excel files and bulk-creating Corpus objects with timing stats.
"""
from typing import Tuple, Iterable, Generator
import openpyxl
import time
import statistics
import logging
from django.db import transaction
from .models import Corpus

logger = logging.getLogger(__name__)


def parse_excel_file(file_obj) -> Tuple[list, Generator]:
    """Parse an uploaded Excel file in read-only mode.

    Returns (headers, rows_generator). The rows_generator yields row tuples (values_only=True) starting from row 2.
    The caller must consume the generator; the workbook will be closed when generator finishes.
    Raises ValueError if header row is missing or malformed.
    """
    workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active

    # Read header
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row or len(first_row) < 1:
        workbook.close()
        raise ValueError('Excel 文件格式不正确，至少需要一列数据')

    def rows_gen():
        try:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                yield row
        finally:
            # Ensure workbook is closed once generator is exhausted or on error
            try:
                workbook.close()
            except Exception:
                pass

    return list(first_row), rows_gen()


def bulk_create_corpora(evaluation_set, rows: Iterable, batch_size: int = 1000, logger_obj: logging.Logger = None) -> dict:
    """Bulk create Corpus objects from an iterable of rows.

    Each row is expected to be a sequence where index 0 is content, 1 is expected_response (optional), 2 is intent (optional).
    Returns a stats dict: {success_count, total_time, rows_per_sec, batch_stats}
    """
    if logger_obj is None:
        logger_obj = logger

    start_time = time.time()
    objs = []
    success_count = 0
    batch_durations = []

    with transaction.atomic():
        for row in rows:
            if row and row[0]:
                content = str(row[0])
                expected_response = str(row[1]) if len(row) > 1 and row[1] is not None else None
                intent = str(row[2]) if len(row) > 2 and row[2] is not None else None

                objs.append(Corpus(
                    evaluation_set=evaluation_set,
                    content=content,
                    expected_response=expected_response,
                    intent=intent
                ))

                if len(objs) >= batch_size:
                    t0 = time.time()
                    Corpus.objects.bulk_create(objs, batch_size=batch_size)
                    elapsed = time.time() - t0
                    batch_durations.append(elapsed)
                    success_count += len(objs)
                    logger_obj.debug('bulk_create: inserted %d rows in %.3fs', len(objs), elapsed)
                    objs = []

        # final
        if objs:
            t0 = time.time()
            Corpus.objects.bulk_create(objs, batch_size=batch_size)
            elapsed = time.time() - t0
            batch_durations.append(elapsed)
            success_count += len(objs)
            logger_obj.debug('bulk_create: inserted final %d rows in %.3fs', len(objs), elapsed)

    total_time = time.time() - start_time
    rows_per_sec = success_count / total_time if total_time > 0 else 0

    try:
        batch_stats = {
            'batches': len(batch_durations),
            'batch_min': min(batch_durations) if batch_durations else 0,
            'batch_max': max(batch_durations) if batch_durations else 0,
            'batch_mean': statistics.mean(batch_durations) if batch_durations else 0,
        }
    except Exception:
        batch_stats = {'batches': len(batch_durations)}

    stats = {
        'success_count': success_count,
        'total_time': total_time,
        'rows_per_sec': rows_per_sec,
        'batch_stats': batch_stats,
    }

    logger_obj.info('bulk_create_corpora: inserted %d rows in %.3fs (%.2f rows/s) batches=%s',
                    success_count, total_time, rows_per_sec, batch_stats)

    return stats
